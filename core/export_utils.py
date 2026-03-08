from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


CURRENCY_FORMAT = '#,##0 "so\'m"'


def _excel_response(workbook, filename):
    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(sheet, row=1):
    fill = PatternFill('solid', fgColor='DDEEFF')
    for cell in sheet[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_currency_column(sheet, col_letter, start_row, end_row):
    for r in range(start_row, end_row + 1):
        cell = sheet[f'{col_letter}{r}']
        cell.number_format = CURRENCY_FORMAT
        cell.alignment = Alignment(horizontal='right')


def _currency(value):
    return f"{Decimal(value):,.0f} so'm"


def _signed_currency(value):
    amount = Decimal(value)
    if amount > 0:
        return f"+{abs(amount):,.0f} so'm"
    if amount < 0:
        return f"-{abs(amount):,.0f} so'm"
    return "0 so'm"


def _order_balance_currency(remaining):
    amount = Decimal(remaining)
    if amount > 0:
        return f"-{abs(amount):,.0f} so'm"
    if amount < 0:
        return f"+{abs(amount):,.0f} so'm"
    return "0 so'm"


def _shop_balance_currency(balance):
    amount = Decimal(balance)
    if amount < 0:
        return f"-{abs(amount):,.0f} so'm"
    if amount > 0:
        return f"+{abs(amount):,.0f} so'm"
    return "0 so'm"


def export_orders_excel(orders):
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"
    ws.append(['Buyurtma ID', "Do'kon", 'Sana', 'Jami summa', "To'langan", 'Qoldiq'])
    _style_header(ws)

    row_no = 1
    for order in orders:
        row_no += 1
        ws.append(
            [
                order.id,
                order.shop.name,
                order.order_date.strftime('%d.%m.%Y'),
                float(order.total_amount),
                float(order.paid_amount),
                float(order.remaining_balance),
            ]
        )

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    if row_no > 1:
        _style_currency_column(ws, 'D', 2, row_no)
        _style_currency_column(ws, 'E', 2, row_no)
        _style_currency_column(ws, 'F', 2, row_no)

    return _excel_response(wb, 'buyurtmalar.xlsx')


def export_order_excel(order):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Buyurtma_{order.id}"

    ws.append(['Buyurtma ID', order.id])
    ws.append(["Do'kon", order.shop.name])
    ws.append(['Sana', order.order_date.strftime('%d.%m.%Y')])
    ws.append([])
    ws.append(['Mahsulot', 'Soni', 'Narxi', 'Jami'])
    _style_header(ws, row=5)

    start_items = 6
    row_no = 5
    for item in order.items.select_related('product'):
        row_no += 1
        ws.append([item.product.name, item.quantity, float(item.price_at_sale), float(item.total_amount)])

    ws.append([])
    ws.append(['Umumiy summa', float(order.total_amount)])
    ws.append(["To'langan", float(order.paid_amount)])
    ws.append(['Qoldiq', float(order.remaining_balance)])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    if row_no >= start_items:
        _style_currency_column(ws, 'C', start_items, row_no)
        _style_currency_column(ws, 'D', start_items, row_no)

    summary_start = row_no + 2
    _style_currency_column(ws, 'B', summary_start, summary_start + 2)

    return _excel_response(wb, f'buyurtma_{order.id}.xlsx')


def export_shops_excel(shops):
    wb = Workbook()
    ws = wb.active
    ws.title = "Do'konlar"
    ws.append([
        "Do'kon nomi",
        'Manzil',
        'Telefon 1',
        'Telefon 2',
        'Izoh',
        'Jami olgan',
        "Jami to'lagan",
        'Balans',
    ])
    _style_header(ws)

    row_no = 1
    for shop in shops:
        row_no += 1
        ws.append(
            [
                shop.name,
                shop.address,
                shop.phone_primary,
                shop.phone_secondary,
                shop.note,
                float(shop.total_purchased),
                float(shop.total_paid),
                float(shop.balance),
            ]
        )

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    if row_no > 1:
        _style_currency_column(ws, 'F', 2, row_no)
        _style_currency_column(ws, 'G', 2, row_no)
        _style_currency_column(ws, 'H', 2, row_no)

    return _excel_response(wb, 'dokonlar.xlsx')


def export_transactions_excel(shop, transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tranzaksiyalar"
    ws.append(["Do'kon", shop.name])
    ws.append(['Manzil', shop.address or '-'])
    ws.append(['Telefonlar', f"{shop.phone_primary or '-'} / {shop.phone_secondary or '-'}"])
    ws.append(['Izoh', shop.note or '-'])
    ws.append([])
    ws.append(['Sana', 'Turi', 'Miqdor', 'Buyurtma ID', 'Izoh'])
    _style_header(ws, row=6)

    row_no = 6
    for row in transactions:
        row_no += 1
        ws.append(
            [
                row['date'].strftime('%d.%m.%Y'),
                row['type'],
                float(row['amount']),
                row.get('order_id') or '-',
                row.get('note') or '-',
            ]
        )

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 40

    if row_no > 6:
        _style_currency_column(ws, 'C', 7, row_no)

    return _excel_response(wb, f"dokon_{shop.id}_tranzaksiyalar.xlsx")


def export_analytics_excel(payload):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Analitika'

    ws.append(['Hisobot oyi', payload['selected_month']])
    ws.append(['Bu oy savdo summasi', payload['month_sales']])
    ws.append([])
    ws.append(['Mahsulot', 'Sotilgan dona', 'Tushgan pul'])
    _style_header(ws, row=4)

    row_no = 4
    for row in payload['product_rows']:
        row_no += 1
        ws.append([row['name'], row['quantity'], float(row['revenue'])])

    ws2 = wb.create_sheet('Kunlik trend')
    ws2.append(['Kun', 'Savdo'])
    _style_header(ws2)
    row2 = 1
    for label, value in zip(payload['line_labels'], payload['line_values']):
        row2 += 1
        ws2.append([label, value])

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 16

    if row_no > 4:
        _style_currency_column(ws, 'C', 5, row_no)
    if row2 > 1:
        _style_currency_column(ws2, 'B', 2, row2)

    return _excel_response(wb, f"analitika_{payload['selected_month']}.xlsx")


def _pdf_escape(text):
    return str(text).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _pdf_text(commands, text, x, y, size=10, bold=False, align='left', max_w=0):
    approx_char = size * 0.5
    width = len(str(text)) * approx_char
    tx = x
    if align == 'right' and max_w:
        tx = x + max_w - width - 4
    elif align == 'center' and max_w:
        tx = x + (max_w - width) / 2
    font = '/F2' if bold else '/F1'
    commands.append(f'BT {font} {size} Tf {tx:.2f} {y:.2f} Td ({_pdf_escape(text)}) Tj ET')


def _pdf_rect(commands, x, y, w, h, fill=False):
    if fill:
        commands.append(f'{x:.2f} {y:.2f} {w:.2f} {h:.2f} re f')
    else:
        commands.append(f'{x:.2f} {y:.2f} {w:.2f} {h:.2f} re S')


def _build_professional_pdf(title, subtitle, headers, rows, numeric_cols, summary_lines):
    page_w = 595
    top = 820
    left = 30
    content_w = page_w - 2 * left

    ncols = len(headers)
    col_w = content_w / ncols
    col_widths = [col_w for _ in headers]

    commands = []

    commands.append('0.12 0.35 0.72 rg')
    _pdf_rect(commands, left, top - 8, 130, 20, fill=True)
    commands.append('0 g')
    _pdf_text(commands, 'Suv Savdo Tizimi', left + 8, top - 2, size=10, bold=True)
    _pdf_text(commands, title, left, top - 30, size=15, bold=True)
    _pdf_text(commands, subtitle, left, top - 46, size=10)
    _pdf_text(commands, f"Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}", left + 390, top - 46, size=8)

    y = top - 72

    commands.append('0.93 0.95 0.98 rg')
    _pdf_rect(commands, left, y - 20, content_w, 20, fill=True)
    commands.append('0 g')

    x = left
    for i, head in enumerate(headers):
        _pdf_text(commands, head, x, y - 14, size=9, bold=True, align='center', max_w=col_widths[i])
        x += col_widths[i]

    commands.append('0.75 0.75 0.75 RG')
    commands.append('0.7 w')
    _pdf_rect(commands, left, y - 20, content_w, 20)

    y -= 20

    row_h = 18
    max_rows = 28
    for row in rows[:max_rows]:
        x = left
        for i, val in enumerate(row):
            _pdf_rect(commands, x, y - row_h, col_widths[i], row_h)
            align = 'right' if i in numeric_cols else 'left'
            _pdf_text(commands, val, x + 3, y - 12, size=8.7, align=align, max_w=col_widths[i] - 6)
            x += col_widths[i]
        y -= row_h

    y -= 12
    for line in summary_lines:
        _pdf_text(commands, line, left, y, size=10, bold=True)
        y -= 14

    _pdf_text(commands, 'Hisobot avtomatik shakllantirildi.', left, 32, size=8)

    stream = '\n'.join(commands).encode('latin-1', errors='replace')

    objects = [
        b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n',
        b'2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n',
        (
            b'3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] '
            b'/Resources << /Font << /F1 4 0 R /F2 5 0 R >> >> /Contents 6 0 R >> endobj\n'
        ),
        b'4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n',
        b'5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >> endobj\n',
        b'6 0 obj << /Length ' + str(len(stream)).encode() + b' >> stream\n' + stream + b'\nendstream endobj\n',
    ]

    pdf = b'%PDF-1.4\n'
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj
    xref_offset = len(pdf)
    pdf += b'xref\n0 7\n0000000000 65535 f \n'
    for off in offsets[1:]:
        pdf += f'{off:010d} 00000 n \n'.encode()
    pdf += b'trailer << /Size 7 /Root 1 0 R >>\nstartxref\n' + str(xref_offset).encode() + b'\n%%EOF'
    return pdf


def pdf_response(filename, data):
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_orders_pdf(orders):
    rows = []
    total = Decimal('0')
    paid = Decimal('0')
    for order in orders:
        rows.append(
            [
                f'#{order.id}',
                order.shop.name,
                order.order_date.strftime('%d.%m.%Y'),
                _currency(order.total_amount),
                _currency(order.paid_amount),
                _order_balance_currency(order.remaining_balance),
            ]
        )
        total += order.total_amount
        paid += order.paid_amount

    data = _build_professional_pdf(
        title='Buyurtmalar hisoboti',
        subtitle='Umumiy buyurtmalar ro\'yxati',
        headers=['Order ID', "Do'kon", 'Sana', 'Total', 'Paid', 'Balance'],
        rows=rows,
        numeric_cols={3, 4, 5},
        summary_lines=[
            f"Jami savdo: {_currency(total)}",
            f"Jami to'langan: {_currency(paid)}",
            f"Jami qoldiq: {_order_balance_currency(total - paid)}",
        ],
    )
    return pdf_response('buyurtmalar.pdf', data)


def export_order_pdf(order):
    rows = []
    for item in order.items.select_related('product'):
        rows.append([item.product.name, f'{item.quantity:,}', _currency(item.price_at_sale), _currency(item.total_amount)])

    data = _build_professional_pdf(
        title=f'Buyurtma #{order.id} invoice',
        subtitle=f"Do'kon: {order.shop.name} | Sana: {order.order_date.strftime('%d.%m.%Y')}",
        headers=['Mahsulot', 'Soni', 'Narxi', 'Jami'],
        rows=rows,
        numeric_cols={1, 2, 3},
        summary_lines=[
            f"Umumiy summa: {_currency(order.total_amount)}",
            f"To'langan summa: {_currency(order.paid_amount)}",
            f"Balans: {_order_balance_currency(order.remaining_balance)}",
        ],
    )
    return pdf_response(f'buyurtma_{order.id}.pdf', data)


def export_shops_pdf(shops):
    rows = []
    for shop in shops:
        rows.append(
            [
                shop.name,
                (shop.address or '-')[:24],
                (shop.phone_primary or '-'),
                (shop.phone_secondary or '-'),
                (shop.note or '-')[:20],
                _currency(shop.total_purchased),
                _currency(shop.total_paid),
                _shop_balance_currency(shop.balance),
            ]
        )

    data = _build_professional_pdf(
        title="Do'konlar hisoboti",
        subtitle="Do'konlar bo'yicha umumiy ko'rsatkichlar",
        headers=["Do'kon", 'Manzil', 'Tel-1', 'Tel-2', 'Izoh', 'Purchased', 'Paid', 'Balance'],
        rows=rows,
        numeric_cols={5, 6, 7},
        summary_lines=[f"Jami do'konlar: {len(rows):,}"],
    )
    return pdf_response('dokonlar.pdf', data)


def export_analytics_pdf(payload):
    rows = []
    for row in payload['product_rows']:
        rows.append([row['name'], f"{int(row['quantity']):,}", _currency(row['revenue'])])

    data = _build_professional_pdf(
        title=f"Analitika hisoboti ({payload['selected_month']})",
        subtitle='Mahsulotlar kesimida savdo natijalari',
        headers=['Mahsulot', 'Sotilgan dona', 'Tushgan pul'],
        rows=rows,
        numeric_cols={1, 2},
        summary_lines=[
            f"Bu oy savdo summasi: {_currency(payload['month_sales'])}",
            f"Trend nuqtalari soni: {len(payload['line_values'])}",
        ],
    )
    return pdf_response(f"analitika_{payload['selected_month']}.pdf", data)


def export_transactions_pdf(shop, transactions):
    rows = []
    for row in transactions:
        rows.append(
            [
                row['date'].strftime('%d.%m.%Y'),
                row['type'],
                _signed_currency(row['amount']),
                f"#{row['order_id']}" if row.get('order_id') else '-',
                (row.get('note') or '-')[:26],
            ]
        )

    data = _build_professional_pdf(
        title=f"Do'kon tarixi: {shop.name}",
        subtitle="Buyurtma, to'lov va depozit tranzaksiyalari",
        headers=['Sana', 'Turi', 'Amount', 'Order ID', 'Izoh'],
        rows=rows,
        numeric_cols={2},
        summary_lines=[
            f"Manzil: {shop.address or '-'}",
            f"Telefonlar: {shop.phone_primary or '-'} / {shop.phone_secondary or '-'}",
            f"Izoh: {shop.note or '-'}",
            f"Jami yozuvlar: {len(rows):,}",
            f"Joriy balans: {_shop_balance_currency(shop.balance)}",
        ],
    )
    return pdf_response(f"dokon_{shop.id}_tarix.pdf", data)
